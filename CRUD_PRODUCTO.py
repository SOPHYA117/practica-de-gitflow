from datetime import datetime
from openpyxl import load_workbook
Rut="C:\\Users\\HP\\Documents\\JASLY SENA WORK\\Base_crud.xlsx"
Rut=r"C:\Users\HP\Documents\JASLY SENA WORK\Base_crud.xlsx"


def leer(ruta:str, extraer:str):
    Archivo_Excel = load_workbook(ruta)
    Hoja_datos = Archivo_Excel['Datos de Producto']
    Hoja_datos=Hoja_datos['A2':'F'+ str(Hoja_datos.max_row)]

    info={}

    for i in Hoja_datos:
        if isinstance(i[0].value, int):
            info.setdefault(i[0].value,{'Producto':i[1].value, 'Categoria':i[2].value,
            'Precio':i[3].value, 'Cantidad':i[4].value})

    if not(extraer=='todo'):
        info=filtrar(info,extraer )

    for i in info:
        print('********Producto********')
        print('Id:'+str(i)+'\n'+'Titulo:'+str(info[i]['Producto'])+'\n'+'Categoria:'+ str(info[i]['categoria'])+
        '\n'+'Precio:'+str(info[i]['precio'])
        + '\n'+'Cantidad:'+str(info[i]['cantidad']))
        print()
    return

def filtrar(info:dict,filtro:str):
    aux={}
    for i in info:
        if info[i]['Categoria']==filtro:
            aux.setdefault(i,info[i])
    return aux

def actualizar(ruta:str,identificador:int,datos_actualizados:dict):
    Archivo_Excel = load_workbook(ruta)
    Hoja_datos = Archivo_Excel['Datos de Producto']
    Hoja_datos = Hoja_datos['A2':'F'+str(Hoja_datos.max_row)]
    hoja=Archivo_Excel.active

    Producto=2
    Categoria=3
    Precio=5
    Cantidad=6
    encontro=False
    for i in Hoja_datos:
        if i[0].value==identificador:
            fila=i[0].row
            encontro=True
            for d in datos_actualizados:
                if d=='titulo' and not (datos_actualizados[d]==''):
                    hoja.cell(row=fila, column=Producto).value=datos_actualizados[d]
                elif d=='Producto' and not(datos_actualizados[d]==''):
                    hoja.cell(row=fila, column=Categoria).value=datos_actualizados[d]
                elif  d=='Categoria' and not(datos_actualizados[d]==''):
                    hoja.cell(row=fila, column=Precio).value=datos_actualizados[d]
                elif  d=='Precio' and not(datos_actualizados[d]==''):
                    hoja.cell(row=fila, column=Cantidad).value=datos_actualizados[d]
    Archivo_Excel.save(ruta)
    if encontro==False:
        print('Error: No existe una tarea con ese Id')
        print()
    return
def agregar(ruta:int,datos:dict):
    Archivo_Excel = load_workbook(ruta)
    Hoja_datos = Archivo_Excel['Datos de Producto']
    Hoja_datos=Hoja_datos['A2':'F'+str(Hoja_datos.max_row+1)]
    hoja=Archivo_Excel.active

    producto=2
    categoria=3
    Precio=4
    cantidad=5
    for i in Hoja_datos:

        if not(isinstance(i[0].value,int)):
            identificador=i[0].row
            hoja.cell(row=identificador, column=1).value=identificador-1
            hoja.cell(row=identificador, column=producto).value=datos['Producto']
            hoja.cell(row=identificador, column=categoria).value=datos['Categoria']
            hoja.cell(row=identificador, column=Precio).value=datos['Precio']
            hoja.cell(row=identificador, column=cantidad).value=datos['Cantidad']
            break
    Archivo_Excel.save(ruta)
    return

def borrar(ruta,identificador):
    Archivo_Excel = load_workbook(ruta)
    Hoja_datos = Archivo_Excel['Datos de Producto']
    Hoja_datos=Hoja_datos['A2':'F'+str(Hoja_datos.max_row)]
    hoja=Archivo_Excel.active

    Producto=2
    Categoria=3
    Precio=4
    Cantidad=5
    encontro=False
    for i in Hoja_datos:
        if i[0].value==identificador:
            fila=i[0].row
            encontro=True

            hoja.cell(row=fila, column=1).value=""
            hoja.cell(row=fila, column=Producto).value=""
            hoja.cell(row=fila, column=Categoria).value=""
            hoja.cell(row=fila, column=Precio).value=""
            hoja.cell(row=fila, column=Cantidad).value=""
    Archivo_Excel.save(ruta)
    if encontro==False:
        print('Error: No existe una tarea con ese id')
        print()
    return

Rut="C:\\Users\\HP\\Documents\\JASLY SENA WORK\\Base_crud.xlsx"
datosActualizados={'producto':'','categoria':'','Precio':'','cantidad':''}
while True:
    print('Indique la accion que desea realizar:')
    print('Consultar 1')
    print('Actualizar: 2')
    print('Crear nueva Categoria: 3')
    print('Borrar: 4')
    accion =input('Escriba la accion:')

    if not(accion=='1') and not (accion=='2') and not (accion=='3') and not (accion=='4'):
        print('Comando invalido por favor elija una opcion valida')
    elif accion=='1':
         opc_consulta=''
         print('Indique la Categoria que desea consultar:')
         print('Todas las Categorias: 1')
         print('Evaluando: 2')
         print('En Venta: 3')
         print('Por Comprar: 4')
         print('Agotado: 5')
         opc_consulta = input('Escriba la Categoria que desea consultar:')
         if opc_consulta=='1':
            print()
            print()
            print('**Consultando todas las Categorias **')
            leer(Rut,'todo')
         if opc_consulta=='2':
            print()
            print()
            print('**Consultando Categoria Evaluada **')
            leer(Rut,'Evaluada')
         if opc_consulta=='3':
            print()
            print()
            print('**Consultando  Categoria en Venta **')
            leer(Rut,'En Venta')
         if opc_consulta=='4':
            print()
            print()
            print('**Consultando  Categoria Por Comprar**')
            leer(Rut,'Por Comprar')
         if opc_consulta=='5':
            print()
            print()
            print('**Consultando  Categoria Agotada**')
            leer(Rut,'Agotada*')
    elif accion=='2':
        datosActualizados={'producto':'','categoria':'','Precio':'','cantidad':''}
        print('**Actualizar tarea**')
        print()
        Id_Actualizar=int(input('Indique el Id de la tarea que desea actualizar:'))
        print()
        print('**Nuevo producto **')
        print('**Nota: si no desea actualizar el titulo solo oprima Enter')
        datosActualizados['producto']=input('Indique el nuevo titulo de la tarea :')
        print()
        print('**Nueva Categoria**')
        print('**Nota: si no desea actualizar la Categoria oprima ENTER')
        datosActualizados['producto']=input('Indique la nueva Categoria  de la tarea :')
        print()
        print('**Nueva Categoria**')
        print('Evaluado: 2')
        print('En Venta: 3')
        print('Por Comprar: 4')
        print('Agotado: 5')
        print('**Nota: si no desea actualizar la Categoria solo oprima ENTER')
        estadoNuevo= input('Indique el nuevo estado de la tarea: ')
        if estadoNuevo=='2':
            datosActualizados['Categoria']='Evaluado'
        elif estadoNuevo=='3':
            datosActualizados['Categoria']= 'En Venta'
        elif estadoNuevo=='4':
            datosActualizados['Categoria']= 'Por Comprar'
        elif estadoNuevo=='5':
            now=datetime.now()
            datosActualizados['Categoria']= 'Agotada'    
        now = datetime.now()
        datosActualizados['fecha']=str(now.day) + '/' + str(now.month) + '/' + str(now.year)
        actualizar(Rut,Id_Actualizar, datosActualizados)
        print()
    elif accion=='3':
        datosActualizados={'Producto':'', 'Categoria':'', 'Precio':'', 'Cantidad':''}
        print('** Crear nuevo Producto **')
        print()
        print('** Producto **')
        print()
        datosActualizados['Producto']=input('Indique el Nombre del Producto: ')
        print()
        print('** Categoria **')
        datosActualizados['Categoria']= input('Indique la categoria del Producto : ')
        print()
        datosActualizados['Categoria']='Evaluado'
        print('** Indique El Precio **')
        print()
        print('** Precio **')
        datosActualizados['Precio']=input('indique el precio del producto : ')
        print('** Indique La Cantidad **')
        print()
        print('** Cantidad **')
        datosActualizados['Cantidad']= input('Indique La Cantidad de producto: ')
        now=datetime.now()
        datosActualizados['fecha inicio']=str(now.day)+'/'+str(now.month)+'/'+str(now.year)
        datosActualizados['fecha finalizacion']=''
        agregar(Rut, datosActualizados)
    elif accion=='4':
        print('')
        print('** Eliminar Producto**')
        iden=int(input('Indique el ID del Producto que desea eliminar : '))  
        borrar(Rut,iden)  